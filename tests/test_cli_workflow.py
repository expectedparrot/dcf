import json
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


class CliWorkflowTests(unittest.TestCase):
    def run_dcf(self, tmp: Path, *args: str, check: bool = True):
        env = dict(os.environ)
        env["PYTHONPATH"] = str(ROOT)
        cmd = [sys.executable, "-m", "dcf", "--dir", str(tmp / ".dcf"), *args]
        result = subprocess.run(cmd, cwd=tmp, env=env, text=True, capture_output=True)
        if check and result.returncode != 0:
            self.fail(f"command failed: {' '.join(cmd)}\nstdout={result.stdout}\nstderr={result.stderr}")
        return result

    def json_cmd(self, tmp: Path, *args: str):
        result = self.run_dcf(tmp, *args)
        return json.loads(result.stdout)

    def seed_model(self, tmp: Path):
        self.run_dcf(tmp, "init", "--company", "Acme", "--author", "jane", "--years", "5")
        assumptions = [
            ("revenue.initial", "1000"),
            ("revenue.growth_rate", "0.10"),
            ("margins.ebit", "0.25"),
            ("tax_rate", "0.21"),
            ("capex.pct_revenue", "0.04"),
            ("wacc", "0.09"),
            ("terminal.growth_rate", "0.025"),
            ("net_debt", "100"),
            ("shares_outstanding", "10"),
        ]
        for key, value in assumptions:
            self.run_dcf(tmp, "assume", "set", key, value, "--author", "jane")

    def test_init_assume_run_and_verify(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.seed_model(tmp)

            assumptions = self.json_cmd(tmp, "assume", "list", "--format", "json")
            self.assertEqual(assumptions["revenue.initial"], 1000)
            self.assertEqual(assumptions["wacc"], 0.09)

            run = self.json_cmd(tmp, "run", "--format", "json", "--tag", "base case")
            self.assertEqual(run["id"], "r001")
            self.assertEqual(run["scenario"], "base")
            self.assertGreater(run["outputs"]["enterprise_value"], 0)
            self.assertGreater(run["outputs"]["implied_share_price"], 0)

            verify = self.json_cmd(tmp, "verify", "--format", "json")
            self.assertTrue(verify["clean"])

    def test_scenario_revision_run_and_verify_detects_tampering(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.seed_model(tmp)
            self.run_dcf(tmp, "scenario", "create", "bull", "--from", "base")
            self.run_dcf(tmp, "scenario", "set", "bull", "revenue.growth_rate", "0.15")

            scenarios = self.json_cmd(tmp, "scenario", "list", "--format", "json")
            self.assertIn("bull", {row["name"] for row in scenarios})

            run = self.json_cmd(tmp, "run", "--scenario", "bull", "--format", "json")
            self.assertEqual(run["scenario"], "bull")
            self.assertIsNotNone(run["scenario_seq"])

            assumption_file = next((tmp / ".dcf" / "assumptions").glob("*.json"))
            assumption_file.write_text(assumption_file.read_text() + "\n", encoding="utf-8")
            result = self.run_dcf(tmp, "verify", "--format", "json", check=False)
            self.assertEqual(result.returncode, 2)
            self.assertIn("hash_mismatch", result.stderr)

    def test_proposals_can_compare_run_and_promote(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.seed_model(tmp)
            session = self.json_cmd(
                tmp,
                "proposal",
                "session",
                "create",
                "--participant",
                "jane",
                "--participant",
                "alex",
                "--label",
                "IC cases",
                "--format",
                "json",
            )
            self.assertEqual(session["session_id"], "s001")

            p1 = self.json_cmd(
                tmp,
                "proposal",
                "submit",
                "--session",
                "s001",
                "--participant",
                "jane",
                "--label",
                "downside",
                "--set",
                "revenue.growth_rate=0.05",
                "--set",
                "wacc=0.105",
                "--format",
                "json",
            )
            p2 = self.json_cmd(
                tmp,
                "proposal",
                "submit",
                "--session",
                "s001",
                "--participant",
                "alex",
                "--label",
                "upside",
                "--set",
                "revenue.growth_rate=0.14",
                "--format",
                "json",
            )
            self.assertEqual([p1["proposal_id"], p2["proposal_id"]], ["p001", "p002"])

            comparison = self.json_cmd(tmp, "proposal", "compare", "--session", "s001", "--run", "--format", "json")
            self.assertEqual({row["proposal_id"] for row in comparison}, {"p001", "p002"})
            self.assertTrue(all("implied_share_price" in row for row in comparison))

            promoted = self.json_cmd(
                tmp,
                "proposal",
                "promote",
                "--session",
                "s001",
                "--proposal",
                "p001",
                "--scenario",
                "jane-downside",
                "--format",
                "json",
            )
            self.assertEqual(promoted["provenance"]["source"], "proposal")
            self.assertEqual(promoted["provenance"]["participant_id"], "jane")

            verify = self.json_cmd(tmp, "verify", "--format", "json")
            self.assertTrue(verify["clean"])

    def test_export_excel_scenario_contains_dcf_formulas(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.seed_model(tmp)
            self.run_dcf(tmp, "scenario", "create", "bull", "--from", "base")
            self.run_dcf(tmp, "scenario", "set", "bull", "revenue.growth_rate", "0.15")
            out = tmp / "bull_case.xlsx"

            result = self.run_dcf(tmp, "export", "excel", "--scenario", "bull", "--out", str(out))
            self.assertIn(str(out), result.stdout)
            self.assertTrue(out.exists())

            with zipfile.ZipFile(out) as zf:
                names = set(zf.namelist())
                self.assertIn("xl/workbook.xml", names)
                self.assertIn("xl/worksheets/sheet1.xml", names)
                self.assertIn("xl/worksheets/sheet2.xml", names)
                dcf_sheet = zf.read("xl/worksheets/sheet2.xml").decode("utf-8")
                inputs_sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

            self.assertIn("revenue.initial", inputs_sheet)
            self.assertIn("<f>Inputs!$B$4*(1+B4)</f>", dcf_sheet)
            self.assertIn("<f>B5*(1+C4)</f>", dcf_sheet)
            self.assertIn("<f>B5*B6</f>", dcf_sheet)
            self.assertIn("<f>B9+B11-B14-B17</f>", dcf_sheet)
            self.assertIn("<f>IF(B24&gt;0,F12*B24,F18*(1+$B$23)/($B$22-$B$23))</f>", dcf_sheet)
            self.assertIn("<f>SUM(B20:F20)+B29</f>", dcf_sheet)

            wb = load_workbook(out, data_only=False)
            self.assertEqual(wb.sheetnames, ["Inputs", "DCF", "Audit", "Sensitivity"])
            self.assertEqual(wb["DCF"]["B30"].value, "=SUM(B20:F20)+B29")
            self.assertEqual(wb["DCF"]["B32"].value, "=B31/B26")
            self.assertGreaterEqual(len(wb["DCF"]._charts), 2)
            self.assertTrue(wb["Inputs"].protection.sheet)

    def test_sensitivity_and_scenario_compare_report_deltas(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.seed_model(tmp)
            self.run_dcf(tmp, "scenario", "create", "bull", "--from", "base")
            self.run_dcf(tmp, "scenario", "set", "bull", "revenue.growth_rate", "0.15")

            sensitivity = self.json_cmd(
                tmp,
                "sensitivity",
                "wacc",
                "0.07:0.11:3",
                "terminal.growth_rate",
                "0.01:0.03:3",
                "--scenario",
                "bull",
                "--metric",
                "implied_share_price",
                "--format",
                "json",
            )
            self.assertEqual(sensitivity["values_a"], [0.07, 0.09, 0.11])
            self.assertEqual(sensitivity["values_b"], [0.01, 0.02, 0.03])
            self.assertEqual(len(sensitivity["rows"]), 3)
            self.assertEqual(len(sensitivity["rows"][0]["values"]), 3)

            comparison = self.json_cmd(tmp, "scenario", "compare", "base", "bull", "--run", "--format", "json")
            self.assertEqual(comparison["scenarios"], ["base", "bull"])
            self.assertIn("valuation", comparison)
            self.assertTrue(any(row["key"] == "revenue.growth_rate" for row in comparison["differences"]))
            self.assertIn("delta_vs_first", comparison["valuation"]["bull"])

    def test_parallel_writes_keep_unique_assumption_sequences(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            self.run_dcf(tmp, "init", "--company", "Acme", "--author", "jane")
            env = dict(os.environ)
            env["PYTHONPATH"] = str(ROOT)
            commands = [
                [sys.executable, "-m", "dcf", "--dir", str(tmp / ".dcf"), "assume", "set", "revenue.initial", "1000"],
                [sys.executable, "-m", "dcf", "--dir", str(tmp / ".dcf"), "assume", "set", "wacc", "0.09"],
                [sys.executable, "-m", "dcf", "--dir", str(tmp / ".dcf"), "assume", "set", "tax_rate", "0.21"],
                [sys.executable, "-m", "dcf", "--dir", str(tmp / ".dcf"), "assume", "set", "margins.ebit", "0.25"],
            ]
            procs = [subprocess.Popen(cmd, cwd=tmp, env=env, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE) for cmd in commands]
            for proc in procs:
                stdout, stderr = proc.communicate(timeout=10)
                self.assertEqual(proc.returncode, 0, msg=f"stdout={stdout}\nstderr={stderr}")

            records = [json.loads(path.read_text()) for path in (tmp / ".dcf" / "assumptions").glob("*.json")]
            seqs = sorted(record["seq"] for record in records)
            versions = sorted(record["base_version"] for record in records)
            self.assertEqual(seqs, [1, 2, 3, 4])
            self.assertEqual(versions, [1, 2, 3, 4])


if __name__ == "__main__":
    unittest.main()
