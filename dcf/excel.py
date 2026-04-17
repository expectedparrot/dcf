from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from .core import resolved_scenario, sensitivity_grid_for_inputs, value_series
from .store import DcfRepo, load_json


def input_value(inputs: dict[str, Any], key: str, year_index: int | None = None) -> float:
    value = inputs.get(key, 0)
    if isinstance(value, list):
        index = 0 if year_index is None else min(year_index, len(value) - 1)
        return float(value[index])
    return float(value)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = '$#,##0.00;[Red]($#,##0.00)'
PCT_FMT = "0.0%"
NUMBER_FMT = '#,##0.00'


def write_inputs_sheet(wb: Workbook, inputs: dict[str, Any], projection_years: int) -> None:
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "DCF Inputs"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Assumption"
    ws["B3"] = "Value"

    scalar_keys = ["revenue.initial", "wacc", "terminal.growth_rate", "terminal.exit_multiple", "net_debt", "shares_outstanding"]
    row = 4
    for key in scalar_keys:
        ws.cell(row=row, column=1, value=key)
        value_cell = ws.cell(row=row, column=2, value=input_value(inputs, key))
        value_cell.fill = INPUT_FILL
        value_cell.protection = Protection(locked=False)
        value_cell.number_format = PCT_FMT if key in {"wacc", "terminal.growth_rate"} else NUMBER_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Year")
    for year in range(1, projection_years + 1):
        ws.cell(row=row, column=year + 1, value=year)

    for key in [
        "revenue.growth_rate",
        "margins.ebit",
        "tax_rate",
        "capex.pct_revenue",
        "depreciation.pct_revenue",
        "working_capital.pct_revenue",
    ]:
        row += 1
        ws.cell(row=row, column=1, value=key)
        for idx, value in enumerate(value_series(inputs.get(key, 0), projection_years), start=2):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.fill = INPUT_FILL
            cell.protection = Protection(locked=False)
            cell.number_format = PCT_FMT

    for col in range(1, projection_years + 2):
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=11, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = SUBHEADER_FILL
        ws.cell(row=11, column=col).fill = SUBHEADER_FILL
    ws.column_dimensions["A"].width = 26
    for col in range(2, projection_years + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.protection.sheet = True


def write_dcf_sheet(wb: Workbook, scenario: str, projection_years: int) -> None:
    ws = wb.create_sheet("DCF")
    ws["A1"] = f"DCF - {scenario}"
    ws["A1"].font = Font(bold=True, size=14)
    labels = {
        3: "Year",
        4: "Revenue Growth",
        5: "Revenue",
        6: "EBIT Margin",
        7: "EBIT",
        8: "Tax Rate",
        9: "NOPAT",
        10: "D&A % Revenue",
        11: "D&A",
        12: "EBITDA",
        13: "Capex % Revenue",
        14: "Capex",
        15: "NWC % Revenue",
        16: "Net Working Capital",
        17: "Change in NWC",
        18: "Free Cash Flow",
        19: "Discount Factor",
        20: "PV Free Cash Flow",
        22: "WACC",
        23: "Terminal Growth",
        24: "Terminal Exit Multiple",
        25: "Net Debt",
        26: "Shares Outstanding",
        28: "Terminal Value",
        29: "PV Terminal Value",
        30: "Enterprise Value",
        31: "Equity Value",
        32: "Implied Share Price",
    }
    for row, label in labels.items():
        ws.cell(row=row, column=1, value=label)

    first_col = 2
    last_col = first_col + projection_years - 1
    last_col_letter = get_column_letter(last_col)
    for year in range(1, projection_years + 1):
        col = first_col + year - 1
        letter = get_column_letter(col)
        ws.cell(row=3, column=col, value=year)
        ws.cell(row=4, column=col, value=f"=Inputs!{letter}12")
        if year == 1:
            ws.cell(row=5, column=col, value=f"=Inputs!$B$4*(1+{letter}4)")
        else:
            prev = get_column_letter(col - 1)
            ws.cell(row=5, column=col, value=f"={prev}5*(1+{letter}4)")
        ws.cell(row=6, column=col, value=f"=Inputs!{letter}13")
        ws.cell(row=7, column=col, value=f"={letter}5*{letter}6")
        ws.cell(row=8, column=col, value=f"=Inputs!{letter}14")
        ws.cell(row=9, column=col, value=f"={letter}7*(1-{letter}8)")
        ws.cell(row=10, column=col, value=f"=Inputs!{letter}16")
        ws.cell(row=11, column=col, value=f"={letter}5*{letter}10")
        ws.cell(row=12, column=col, value=f"={letter}7+{letter}11")
        ws.cell(row=13, column=col, value=f"=Inputs!{letter}15")
        ws.cell(row=14, column=col, value=f"={letter}5*{letter}13")
        ws.cell(row=15, column=col, value=f"=Inputs!{letter}17")
        ws.cell(row=16, column=col, value=f"={letter}5*{letter}15")
        if year == 1:
            ws.cell(row=17, column=col, value=f"={letter}16-Inputs!$B$4*{letter}15")
        else:
            prev = get_column_letter(col - 1)
            ws.cell(row=17, column=col, value=f"={letter}16-{prev}16")
        ws.cell(row=18, column=col, value=f"={letter}9+{letter}11-{letter}14-{letter}17")
        ws.cell(row=19, column=col, value=f"=1/(1+$B$22)^{letter}$3")
        ws.cell(row=20, column=col, value=f"={letter}18*{letter}19")

    ws["B22"] = "=Inputs!$B$5"
    ws["B23"] = "=Inputs!$B$6"
    ws["B24"] = "=Inputs!$B$7"
    ws["B25"] = "=Inputs!$B$8"
    ws["B26"] = "=Inputs!$B$9"
    ws["B28"] = f'=IF(B24>0,{last_col_letter}12*B24,{last_col_letter}18*(1+$B$23)/($B$22-$B$23))'
    ws["B29"] = f"=B28/(1+$B$22)^{projection_years}"
    ws["B30"] = f"=SUM(B20:{last_col_letter}20)+B29"
    ws["B31"] = "=B30-B25"
    ws["B32"] = "=B31/B26"

    for row in [3, 22, 28]:
        for col in range(1, projection_years + 2):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
    for row in [30, 31, 32]:
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
    ws.column_dimensions["A"].width = 24
    for col in range(2, projection_years + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(4, 33):
        for col in range(2, projection_years + 2):
            if row in {4, 6, 8, 10, 13, 15, 22, 23}:
                ws.cell(row=row, column=col).number_format = PCT_FMT
            else:
                ws.cell(row=row, column=col).number_format = MONEY_FMT if row != 26 else NUMBER_FMT
    ws.protection.sheet = True


def write_audit_sheet(wb: Workbook, scenario: str, inputs: dict[str, Any]) -> None:
    ws = wb.create_sheet("Audit")
    ws.append(["Scenario", scenario])
    ws.append(["Generated By", "dcf export excel"])
    ws.append([])
    ws.append(["Assumption", "Resolved Value"])
    for key in sorted(inputs):
        ws.append([key, str(inputs[key])])
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def write_sensitivity_sheet(wb: Workbook, inputs: dict[str, Any], projection_years: int) -> None:
    ws = wb.create_sheet("Sensitivity")
    grid = sensitivity_grid_for_inputs(
        inputs,
        projection_years,
        "wacc",
        [0.07, 0.08, 0.09, 0.10, 0.11],
        "terminal.growth_rate",
        [0.01, 0.02, 0.03, 0.04],
        "implied_share_price",
    )
    ws["A1"] = "Implied Share Price Sensitivity"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "WACC \\ Terminal Growth"
    for col, value in enumerate(grid["values_b"], start=2):
        ws.cell(row=3, column=col, value=value).number_format = PCT_FMT
    for row_idx, row in enumerate(grid["rows"], start=4):
        ws.cell(row=row_idx, column=1, value=row["value"]).number_format = PCT_FMT
        for col_idx, value in enumerate(row["values"], start=2):
            ws.cell(row=row_idx, column=col_idx, value=value).number_format = MONEY_FMT
    for col in range(1, 2 + len(grid["values_b"])):
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = SUBHEADER_FILL
    ws.column_dimensions["A"].width = 24
    for col in range(2, 2 + len(grid["values_b"])):
        ws.column_dimensions[get_column_letter(col)].width = 15


def add_charts(wb: Workbook, projection_years: int) -> None:
    ws = wb["DCF"]
    fcf_chart = LineChart()
    fcf_chart.title = "Free Cash Flow"
    fcf_chart.y_axis.title = "FCF"
    fcf_chart.x_axis.title = "Year"
    data = Reference(ws, min_col=2, max_col=projection_years + 1, min_row=18, max_row=18)
    cats = Reference(ws, min_col=2, max_col=projection_years + 1, min_row=3)
    fcf_chart.add_data(data, from_rows=True, titles_from_data=False)
    fcf_chart.set_categories(cats)
    ws.add_chart(fcf_chart, "D4")

    summary = BarChart()
    summary.title = "Valuation Summary"
    data = Reference(ws, min_col=2, min_row=30, max_row=32)
    cats = Reference(ws, min_col=1, min_row=30, max_row=32)
    summary.add_data(data, titles_from_data=False)
    summary.set_categories(cats)
    ws.add_chart(summary, "D20")


def export_excel(repo: DcfRepo, scenario: str, out: str | None) -> Path:
    repo.ensure_exists()
    config = load_json(repo.root / "config.json")
    projection_years = int(config.get("projection_years", 10))
    inputs, _ = resolved_scenario(repo, scenario)
    output = Path(out) if out else Path.cwd() / f"{scenario}_dcf.xlsx"
    if output.is_dir():
        output = output / f"{scenario}_dcf.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    write_inputs_sheet(wb, inputs, projection_years)
    write_dcf_sheet(wb, scenario, projection_years)
    write_audit_sheet(wb, scenario, inputs)
    write_sensitivity_sheet(wb, inputs, projection_years)
    add_charts(wb, projection_years)
    wb.save(output)
    return output
